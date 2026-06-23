import openpyxl

def safe_float(val):
    if val is None:
        return 0.0
    val_str = str(val).strip().upper()
    if val_str in ['AB', 'ABS', 'ABSENT', 'A', 'None', 'N/A', '']:
        return 0.0
    try:
        return float(val)
    except ValueError:
        return 0.0

def read_input_excel(file_path):
    """
    Parses the input Excel sheet.
    Assumes that:
    - Headers are on Row 1 (e.g. Sl. No., Reg. No., Name, Int (40), etc.)
    - Student data starts on Row 2
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    students = []
    # Data starts from row 2 (row 1 contains headers)
    for r in range(2, sheet.max_row + 1):
        s_no_cell = sheet.cell(row=r, column=1).value
        reg_no_cell = sheet.cell(row=r, column=2).value
        name_cell = sheet.cell(row=r, column=3).value
        
        # If Reg. No. and Name are empty, treat as end of student list
        if not reg_no_cell and not name_cell:
            continue
            
        int_40_val = safe_float(sheet.cell(row=r, column=4).value)
        comp_val = int(round(int_40_val / 4.0))
        
        students.append({
            's_no': s_no_cell,
            'reg_no': str(reg_no_cell).strip() if reg_no_cell else '',
            'name': str(name_cell).strip() if name_cell else '',
            'int_40': int_40_val,
            'viva_10': comp_val,
            'day_to_day_10': comp_val,
            'int_10': comp_val,
            'lab_project_10': comp_val,
            'end_sem_60': safe_float(sheet.cell(row=r, column=5).value)
        })
        
    return students

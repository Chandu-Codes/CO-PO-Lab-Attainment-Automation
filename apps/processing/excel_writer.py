import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils import get_column_letter

def get_thin_border():
    thin = Side(border_style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def get_medium_bottom_border():
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=medium)

def get_medium_border_row():
    # Medium border top/bottom for summary cells
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    return Border(left=thin, right=thin, top=medium, bottom=medium)

def write_output_excel(students, config, output_path):
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: Student Marks and CO Attainment Details
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "1"
    
    # Set Grid lines to True
    ws1.views.sheetView[0].showGridLines = True
    
    # 1. Page Merges & Banners
    ws1.merge_cells("A1:S1")
    ws1.merge_cells("A2:S2")
    ws1["A1"] = "MALLA REDDY ENGINEERING COLLEGE  (Autonomous)"
    ws1["A2"] = "B.Tech. - EEE"
    
    ws1["A1"].font = Font(name="Times New Roman", size=12, bold=True)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A2"].font = Font(name="Times New Roman", size=12, bold=True)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Metadata Rows 3 to 6
    metadata = [
        # Row 3
        ("A3:B3", "Year Admitted :", "C3", config.year_admitted, "N3:O3", "Faculty Name :", "P3:Q3", config.faculty_name),
        # Row 4
        ("A4:B4", "Course Code :", "C4", config.course_code, "O4", "Class : ", "P4:R4", config.class_name),
        # Row 5
        ("A5:B5", "Course Name :", "C5", config.course_name, "O5", "Strength: ", "P5:Q5", float(len(students))),
        # Row 6
        ("A6:B6", "Academic Year :", "C6", config.academic_year, "N6:O6", "Regulations :", "P6:Q6", config.regulation)
    ]
    
    for row_info in metadata:
        if len(row_info) == 8:
            m1_range, m1_val, c1_coord, c1_val, m2_range, m2_val, c2_range, c2_val = row_info
            
            # Left block
            ws1.merge_cells(m1_range)
            ws1[m1_range.split(":")[0]] = m1_val
            ws1[c1_coord] = c1_val
            
            # Right block
            ws1.merge_cells(m2_range)
            ws1[m2_range.split(":")[0]] = m2_val
            ws1.merge_cells(c2_range)
            ws1[c2_range.split(":")[0]] = c2_val
        else:
            # Row 4 & 5 middle configurations
            m1_range, m1_val, c1_coord, c1_val, sing_coord, sing_val, c2_range, c2_val = row_info
            ws1.merge_cells(m1_range)
            ws1[m1_range.split(":")[0]] = m1_val
            ws1[c1_coord] = c1_val
            
            ws1[sing_coord] = sing_val
            ws1.merge_cells(c2_range)
            ws1[c2_range.split(":")[0]] = c2_val
            
    # Apply Font & Alignment to metadata
    for r in range(3, 7):
        for c in range(1, 20):
            cell = ws1.cell(row=r, column=c)
            cell.font = Font(name="Times New Roman", size=11, bold=False)
            if c in [1, 14, 15]:  # Label column headers
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Empty merged row 7 bottom line separator
    ws1.merge_cells("A7:N7")
    thin_bottom = Border(bottom=Side(style="thin", color="000000"))
    for c in range(1, 20):
        ws1.cell(row=7, column=c).border = thin_bottom
        
    # 3. Table Headers Row 8
    headers = [
        "Sl. No.", "Reg. No.", "Name", "Int (40)", "Viva  (10)", 
        "Dat to Day  (10)", "Int (10)", "Lab Project  (10)", "End sem Exam     (60)", 
        "CO 1", "AW", "CO 2", "AW", "CO 3", "AW", "CO 4", "AW", "CO 5", "AW"
    ]
    
    for idx, h in enumerate(headers):
        cell = ws1.cell(row=8, column=idx+1)
        cell.value = h
        cell.font = Font(name="Times New Roman", size=10, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_thin_border()
        
    # 4. Populate Student Data starting from Row 9
    N = len(students)
    for idx, s in enumerate(students):
        r = 9 + idx
        ws1.cell(row=r, column=1, value=float(idx+1)) # Sl. No.
        ws1.cell(row=r, column=2, value=s['reg_no'])
        ws1.cell(row=r, column=3, value=s['name'])
        ws1.cell(row=r, column=4, value=s['int_40'])
        ws1.cell(row=r, column=5, value=f"=ROUND(D{r}/4,0)")
        ws1.cell(row=r, column=6, value=f"=ROUND(D{r}/4,0)")
        ws1.cell(row=r, column=7, value=f"=ROUND(D{r}/4,0)")
        ws1.cell(row=r, column=8, value=f"=ROUND(D{r}/4,0)")
        ws1.cell(row=r, column=9, value=s['end_sem_60'])
        
        # Inject dynamic Excel Formulas for CO Marks and Attainment Levels (AW)
        # Col J (CO 1 Marks): E + F + G + H + I
        ws1.cell(row=r, column=10, value=f"=ROUND((E{r}+F{r}+G{r}+H{r}+I{r}),0)")
        # Col K (CO 1 Level): IF J >= 80 -> 3, ELSE IF J >= 60 -> 2, ELSE 1
        ws1.cell(row=r, column=11, value=f"=IF((J{r}>=80),3,IF(AND(J{r}<80,J{r}>=60),2,IF(J{r}<60,1)))")
        
        # Col L/M (CO 2 Marks/Level)
        ws1.cell(row=r, column=12, value=f"=ROUND((E{r}+F{r}+G{r}+H{r}+I{r}),0)")
        ws1.cell(row=r, column=13, value=f"=IF((L{r}>=80),3,IF(AND(L{r}<80,L{r}>=60),2,IF(L{r}<60,1)))")
        
        # Col N/O (CO 3 Marks/Level)
        ws1.cell(row=r, column=14, value=f"=ROUND((E{r}+F{r}+G{r}+H{r}+I{r}),0)")
        ws1.cell(row=r, column=15, value=f"=IF((N{r}>=80),3,IF(AND(N{r}<80,N{r}>=60),2,IF(N{r}<60,1)))")
        
        # Col P/Q (CO 4 Marks/Level)
        ws1.cell(row=r, column=16, value=f"=ROUND((E{r}+F{r}+G{r}+H{r}+I{r}),0)")
        ws1.cell(row=r, column=17, value=f"=IF((P{r}>=80),3,IF(AND(P{r}<80,P{r}>=60),2,IF(P{r}<60,1)))")
        
        # Col R/S (CO 5 Marks/Level)
        ws1.cell(row=r, column=18, value=f"=ROUND((E{r}+F{r}+G{r}+H{r}+I{r}),0)")
        ws1.cell(row=r, column=19, value=f"=IF((R{r}>=80),3,IF(AND(R{r}<80,R{r}>=60),2,IF(R{r}<60,1)))")
        
        # Stylize student rows
        for c in range(1, 20):
            cell = ws1.cell(row=r, column=c)
            cell.font = Font(name="Times New Roman", size=10)
            cell.border = get_thin_border()
            if c == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # 5. Summary Averages & Attainments
    r_empty1 = 9 + N
    r_empty2 = 10 + N
    r_direct = 11 + N # Row 77 for 66 students
    
    # Direct Attainments calculations in row 11 + N
    for col_idx, col_letter in [(10, "J"), (11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O"), (16, "P"), (17, "Q"), (18, "R"), (19, "S")]:
        cell = ws1.cell(row=r_direct, column=col_idx)
        cell.value = f"=SUM({col_letter}9:{col_letter}{8+N})/P5"
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_thin_border()
        
    r_empty3 = 12 + N
    r_empty4 = 13 + N
    r_empty5 = 14 + N
    r_survey = 15 + N # Row 81 for 66 students
    
    # Write Course End Survey title
    ws1.merge_cells(start_row=r_survey, start_column=5, end_row=r_survey, end_column=8)
    survey_lbl = ws1.cell(row=r_survey, column=5)
    survey_lbl.value = "Course End Survey"
    survey_lbl.font = Font(name="Times New Roman", size=10, bold=True)
    survey_lbl.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write survey formulas from DB configuration to columns K, M, O, Q, S
    survey_data = config.survey_inputs
    ws1.cell(row=r_survey, column=11, value=f"={survey_data.get('CO1', '0')}")
    ws1.cell(row=r_survey, column=13, value=f"={survey_data.get('CO2', '0')}")
    ws1.cell(row=r_survey, column=15, value=f"={survey_data.get('CO3', '0')}")
    ws1.cell(row=r_survey, column=17, value=f"={survey_data.get('CO4', '0')}")
    ws1.cell(row=r_survey, column=19, value=f"={survey_data.get('CO5', '0')}")
    
    # Style survey row elements
    for c in [11, 13, 15, 17, 19]:
        cell = ws1.cell(row=r_survey, column=c)
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    r_empty6 = 16 + N
    r_attain = 17 + N # Row 83 for 66 students
    
    # Combined Attainment (80% Direct + 20% Indirect)
    ws1.merge_cells(start_row=r_attain, start_column=3, end_row=r_attain, end_column=8)
    attain_lbl = ws1.cell(row=r_attain, column=3)
    attain_lbl.value = "CO ATTAINMENT (80% DIRECT AND 20% INDIRECT)"
    attain_lbl.font = Font(name="Times New Roman", size=10, bold=True)
    attain_lbl.alignment = Alignment(horizontal="center", vertical="center")
    
    # Formulas for final combined scores: e.g. =0.8*K77+0.2*K81
    ws1.cell(row=r_attain, column=11, value=f"=0.8*K{r_direct}+0.2*K{r_survey}")
    ws1.cell(row=r_attain, column=13, value=f"=0.8*M{r_direct}+0.2*M{r_survey}")
    ws1.cell(row=r_attain, column=15, value=f"=0.8*O{r_direct}+0.2*O{r_survey}")
    ws1.cell(row=r_attain, column=17, value=f"=0.8*Q{r_direct}+0.2*Q{r_survey}")
    ws1.cell(row=r_attain, column=19, value=f"=0.8*S{r_direct}+0.2*S{r_survey}")
    
    for c in [11, 13, 15, 17, 19]:
        cell = ws1.cell(row=r_attain, column=c)
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Signature block at 20 + N (Row 86)
    r_sigs = 20 + N
    ws1.cell(row=r_sigs, column=3, value="Coordinator").font = Font(name="Times New Roman", size=11, bold=True)
    ws1.cell(row=r_sigs, column=18, value="HOD").font = Font(name="Times New Roman", size=11, bold=True)
    
    # Set column widths for Sheet 1
    widths = {
        'A': 8, 'B': 15, 'C': 35, 'D': 10, 'E': 10, 'F': 12, 'G': 10, 'H': 12, 'I': 15,
        'J': 10, 'K': 8, 'L': 10, 'M': 8, 'N': 10, 'O': 8, 'P': 10, 'Q': 8, 'R': 10, 'S': 8
    }
    for col_let, w in widths.items():
        ws1.column_dimensions[col_let].width = w

    # ----------------------------------------------------
    # SHEET 2: Assessment Table & Graph
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Assessment Table & Graph")
    ws2.views.sheetView[0].showGridLines = True
    
    # Banners rows 1 to 3
    ws2.merge_cells("A1:S1")
    ws2.merge_cells("A2:S2")
    ws2.merge_cells("A3:R3")
    
    ws2["A1"] = "MALLA REDDY ENGINEERING COLLEGE  (Autonomous)"
    ws2["A2"] = "B.Tech. - EEE"
    ws2["A3"] = "Level of Attainment of Course Outcome"
    
    ws2["A1"].font = Font(name="Times New Roman", size=12, bold=True)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A2"].font = Font(name="Times New Roman", size=12, bold=True)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A3"].font = Font(name="Cambria", size=11, bold=True)
    ws2["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata block rows 4-6 referencing Sheet 1
    ws2.merge_cells("B4:D4")
    ws2["B4"] = "Academic Year :"
    ws2.merge_cells("E4:F4")
    ws2["E4"] = "='1'!C3"
    ws2.merge_cells("H4:K4")
    ws2["H4"] = "Faculty Name"
    ws2.merge_cells("L4:N4")
    ws2["L4"] = "='1'!P3"
    
    ws2.merge_cells("B5:D5")
    ws2["B5"] = "Course code :"
    ws2["E5"] = "='1'!C4"
    ws2.merge_cells("H5:K5")
    ws2["H5"] = "Course Name:"
    ws2.merge_cells("L5:N5")
    ws2["L5"] = "='1'!C5"
    
    ws2.merge_cells("B6:D6")
    ws2["B6"] = "Student Strength :"
    ws2["E6"] = "='1'!P5"
    ws2.merge_cells("H6:K6")
    ws2["H6"] = "Regulation"
    ws2.merge_cells("L6:N6")
    ws2["L6"] = config.regulation
    
    # Apply Font to sheet 2 metadata rows
    for r in range(4, 7):
        for c in range(1, 20):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name="Times New Roman", size=10)
            if c in [2, 8]: # labels
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 6. Attainment Counts Grid (Rows 8-15)
    ws2.merge_cells("C8:E10")
    ws2["C8"] = "Course Outcomes"
    ws2.merge_cells("F8:K8")
    ws2["F8"] = "Attainment of COs"
    
    ws2.merge_cells("F9:G9")
    ws2["F9"] = "Level 3"
    ws2.merge_cells("H9:I9")
    ws2["H9"] = "Level 2"
    ws2.merge_cells("J9:K9")
    ws2["J9"] = "Level 1"
    
    ws2["F10"] = "Number"
    ws2["G10"] = "%"
    ws2["H10"] = "Number"
    ws2["I10"] = "%"
    ws2["J10"] = "Number"
    ws2["K10"] = "%"
    
    # Style grid headers
    for r in range(8, 11):
        for c in range(3, 12):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name="Cambria", size=10, bold=(r == 8 and c == 6))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = get_thin_border()
            
    # Inject CO1 to CO5 counts & percentages
    # Let's map sheet 1 columns to outcomes: J (CO1), L (CO2), N (CO3), P (CO4), R (CO5)
    co_mappings = [
        ("CO 1", "J"),
        ("CO 2", "L"),
        ("CO 3", "N"),
        ("CO 4", "P"),
        ("CO 5", "R")
    ]
    
    for idx, (co_lbl, col_let) in enumerate(co_mappings):
        r = 11 + idx
        ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws2.cell(row=r, column=3, value=co_lbl)
        
        # COUNTIF Level 3 (>=80 marks)
        ws2.cell(row=r, column=6, value=f"=COUNTIF('1'!{col_let}9:{col_let}{8+N}, \">=80\")")
        # % Level 3
        ws2.cell(row=r, column=7, value=f"=(F{r}/E6)*100")
        
        # COUNTIF Level 2 (>=60 and <80)
        ws2.cell(row=r, column=8, value=f"=COUNTIF('1'!{col_let}9:{col_let}{8+N}, \">=60\")-COUNTIF('1'!{col_let}9:{col_let}{8+N}, \">=80\")")
        # % Level 2
        ws2.cell(row=r, column=9, value=f"=(H{r}/E6)*100")
        
        # COUNTIF Level 1 (<60)
        ws2.cell(row=r, column=10, value=f"=COUNTIF('1'!{col_let}9:{col_let}{8+N}, \"<60\")")
        # % Level 1
        ws2.cell(row=r, column=11, value=f"=(J{r}/E6)*100")
        
        # Style rows 11 to 15
        for c in range(3, 12):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name="Cambria", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = get_thin_border()

    # 7. Inject Level Distribution BarChart at Row 17 (Cell D17)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.width = 16
    chart.height = 6.8
    
    # Position the legend cleanly to prevent overlap
    chart.legend.position = "r"
    chart.legend.overlay = False
    
    # Adding three series (Level 3 %, Level 2 %, Level 1 %)
    # Values references are Column G (col=7), Column I (col=9), Column K (col=11) in rows 11 to 15
    for col_idx, title in [(7, "Level 3"), (9, "Level 2"), (11, "Level 1")]:
        ref = Reference(ws2, min_col=col_idx, min_row=11, max_row=15)
        series = Series(ref, title=title)
        chart.series.append(series)
        
    # Categories reference is the CO names: C11:C15 (Col 3)
    cats = Reference(ws2, min_col=3, min_row=11, max_row=15)
    chart.set_categories(cats)
    
    ws2.add_chart(chart, "D17")
    
    # 8. PO-Mapping Matrix (Rows 32-40)
    ws2.merge_cells("A32:A33")
    ws2["A32"] = "Name of the course"
    ws2.merge_cells("B32:B33")
    ws2["B32"] = "CO'S"
    ws2.merge_cells("C32:C33")
    ws2["C32"] = "Attained values"
    ws2.merge_cells("D32:S32")
    ws2["D32"] = "Contribution to Program Outcomes"
    
    po_headers = [
        "PO-1", "PO-2", "PO-3", "PO-4", "PO-5", "PO-6", "PO-7", "PO-8", "PO-9", "PO-10", "PO-11", "PO-12", "PSO-1", "PSO-2", "PSO-3", "PSO-3"
    ]
    for idx, ph in enumerate(po_headers):
        ws2.cell(row=33, column=4+idx, value=ph)
        
    # Apply Font to PO matrix headers (Rows 32-33)
    for r in [32, 33]:
        for c in range(1, 20):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name="Times New Roman", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = get_thin_border()
            
    # Rows 34 to 38: Mapped weights for CO1 to CO5
    # Name of course merged A34:A39
    ws2.merge_cells("A34:A39")
    ws2["A34"] = f"{config.course_name} ({config.course_code})"
    
    # Weights mappings
    weights_dict = config.co_po_mapping
    co_list = ["CO1", "CO2", "CO3", "CO4", "CO5"]
    
    for idx, co_key in enumerate(co_list):
        r = 34 + idx
        # Replicate CO3 mislabel for row 37 (idx=3, CO4 is mislabelled as CO3 in target ATG sheet)
        co_label = "CO3" if idx == 3 else f"CO{idx+1}"
        ws2.cell(row=r, column=2, value=co_label)
        
        # Link to Sheet 1 Combined Attainments (Row 17 + N)
        # K (CO1), M (CO2), O (CO3), Q (CO4), S (CO5)
        sheet1_col = ["K", "M", "O", "Q", "S"][idx]
        ws2.cell(row=r, column=3, value=f"='1'!{sheet1_col}{r_attain}")
        
        # Write PO weights
        row_weights = weights_dict.get(co_key, {})
        for w_idx, ph in enumerate(po_headers):
            # Resolve duplicate PSO-3 header representation (Col R and Col S)
            # Col 18 (w_idx=14) is PSO-3, Col 19 (w_idx=15) is empty PSO-3 in target sheet
            if w_idx == 15:
                val = None
            else:
                val = row_weights.get(ph, None)
            # Write weight value to cell
            ws2.cell(row=r, column=4+w_idx, value=float(val) if val is not None else None)
            
        # Style row cells
        for c in range(1, 20):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = get_thin_border()
            
    # Row 39 duplicated CO5 block to replicate exact teacher sheet structure!
    # Wait, the user wants the exact file, let's write it to ensure perfect layout mapping!
    ws2.cell(row=39, column=2, value="CO5")
    # Replicate exact external/broken link from teacher's sheet
    ws2.cell(row=39, column=3, value="='[1]1'!BM84")
    for w_idx, ph in enumerate(po_headers):
        if w_idx == 15:
            val = None
        else:
            val = weights_dict.get("CO5", {}).get(ph, None)
        ws2.cell(row=39, column=4+w_idx, value=float(val) if val is not None else None)
        
    for c in range(1, 20):
        cell = ws2.cell(row=39, column=c)
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_medium_bottom_border()

    # Row 40 (Attainment Values):
    ws2.merge_cells("A40:C40")
    ws2["A40"] = "Attainment Values"
    ws2["A40"].font = Font(name="Times New Roman", size=10, bold=True)
    ws2["A40"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A40"].border = get_medium_border_row()
    
    # Dynamic formula: =IF(SUM(D34:D39)=0,0,(SUM(C34*D34)+(C35*D35)+(C36*D36)+(C38*D38)+(C39*D39))/SUM(D34:D39))
    # Note that Column S (c_idx = 19) uses E instead of C in the numerator!
    for c_idx in range(4, 20):
        col_let = get_column_letter(c_idx)
        cell = ws2.cell(row=40, column=c_idx)
        num_let = "E" if col_let == "S" else "C"
        cell.value = f"=IF(SUM({col_let}34:{col_let}39)=0,0,(SUM({num_let}34*{col_let}34)+({num_let}35*{col_let}35)+({num_let}36*{col_let}36)+({num_let}38*{col_let}38)+({num_let}39*{col_let}39))/SUM({col_let}34:{col_let}39))"
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_medium_border_row()
        
    # Row 43 (Signatures)
    ws2.cell(row=43, column=2, value="Coordinator").font = Font(name="Times New Roman", size=11, bold=True)
    ws2.cell(row=43, column=16, value="HOD").font = Font(name="Times New Roman", size=11, bold=True)

    # Set column widths for Sheet 2
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 15
    for c_idx in range(4, 20):
        ws2.column_dimensions[get_column_letter(c_idx)].width = 8
        
    # Save Workbook
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)
